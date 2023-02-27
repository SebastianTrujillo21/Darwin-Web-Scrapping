import time
import tkinter
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select


def menu_registro_usuario():
    global nombre_usuario_entry
    global contrasena_usuario_entry
    global nombre_usuario_verify
    global contrasena_usuario_verify
    global dirExcel
    global pantalla
    pantalla = tkinter.Tk()
    pantalla.geometry("300x380")
    pantalla.title("Inicio de Sesion")
    tkinter.Label(pantalla, text="").pack()
    image = tkinter.PhotoImage(file="GSI.gif")
    image = image.subsample(2, 2)
    tkinter.Label(pantalla, image=image).pack()
    tkinter.Label(pantalla, text="").pack()

    nombre_usuario_verify = tkinter.StringVar()
    contrasena_usuario_verify = tkinter.StringVar()
    dirExcel = tkinter.StringVar()

    tkinter.Label(pantalla, text="Username").pack()
    nombre_usuario_entry = tkinter.Entry(pantalla, textvariable=nombre_usuario_verify)
    nombre_usuario_entry.pack()
    tkinter.Label(pantalla, text="").pack()

    tkinter.Label(pantalla, text="Password").pack()
    contrasena_usuario_entry = tkinter.Entry(pantalla, textvariable=contrasena_usuario_verify)
    contrasena_usuario_entry.pack()
    tkinter.Label(pantalla, text="").pack()

    tkinter.Button(pantalla, text="Iniciar Sesion", command=abrirArchivo).pack()
    pantalla.mainloop()


def abrirArchivo():
    archivo = filedialog.askopenfilename(initialdir="/", title="Seleccionar el archivo",
                                         filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
    dirExcel.set(archivo)
    pantalla.destroy()


menu_registro_usuario()
username = nombre_usuario_verify.get()  # Usuario: 30030644
password = contrasena_usuario_verify.get()  # Clave: 30644feb

filename = str(dirExcel.get())
filesheet = pd.read_excel(filename, usecols='A')
workbook = load_workbook(filename)
worksheet = workbook.active

dniArray = list(filesheet.loc[:, 'DNI'])
nombreArray = []
primerApellidoArray = []
segundoApellidoArray = []
correoArray = []
telefonoArray = []
cuentaBancariaArray = []
cupsLuzArray = []
cupsGasArray = []
direccionArray = []


def tarea():
    driver.find_element(By.XPATH, "/html/body/div[4]/div[3]/div/div/div[2]/div/div/button").click()
    time.sleep(10)
    driver.find_element(By.XPATH, "/html/body/div[1]/section/div[2]/div/div/button").click()
    time.sleep(10)
    driver.find_element(By.XPATH,
                        '//*[@id="auth0-lock-container-1"]/div/div[2]/form/div/div/div/div/div[2]/div[2]/span/div/div/div/div/div/div/div/div/a/div[2]').click()
    time.sleep(10)
    driver.find_element(By.XPATH, "/html/body/div/div/div/div[2]/div[1]/form/input[3]").send_keys(username)
    driver.find_element(By.XPATH, "/html/body/div/div/div/div[2]/div[1]/form/input[4]").send_keys(password)
    time.sleep(10)
    driver.find_element(By.XPATH, "/html/body/div/div/div/div[2]/div[1]/form/input[5]").click()
    time.sleep(10)
    try:
        driver.find_element(By.XPATH, "//input[@value='Continue']").click()
    except NoSuchElementException:
        print("error")
    time.sleep(30)
    canal = driver.find_element(By.XPATH, "/html/body/div[1]/section/div[2]/div/div/div/div[1]/div/div/div/select")
    selector_canal = Select(canal)
    selector_canal.select_by_visible_text("Fuerza de Ventas con Verificacion")
    time.sleep(10)
    driver.find_element(By.XPATH, "/html/body/div[1]/section/div[2]/div/div/div/div[3]/button").click()
    time.sleep(10)
    for i in range(len(dniArray)):
        time.sleep(5)
        try:
            driver.find_element(By.XPATH,
                                "/html/body/div[1]/section/div[2]/div/div/div[3]/div[1]/div[2]/div[1]/div/div/div/div/form/div[1]/div/div[2]/div/div/input").send_keys(
                dniArray[i])
        except NoSuchElementException:
            driver.refresh()
        time.sleep(10)

        try:
            driver.find_element(By.XPATH,
                                "/html/body/div[1]/section/div[2]/div/div/div[3]/div[1]/div[2]/div[1]/div/div/div/div/form/div[1]/div/div[3]/button").click()
        except ElementClickInterceptedException:
            driver.refresh()
        time.sleep(10)

        try:
            driver.find_element(By.XPATH,
                                "/html/body/div[6]/div[3]/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/div/div/div/div/div/div/div/div[1]/span/span/input").click()
            texto_largo = driver.find_element(By.XPATH, "//div[@columns='[object Object]']").text
            textoLargoArray = texto_largo.split("\n")
            direccionArray.append(textoLargoArray[0])
            for j, contenido in enumerate(textoLargoArray):
                tamanoArrayCupLuz = len(cupsLuzArray)

                if textoLargoArray[j] == "CUPS LUZ":
                    cupsLuzArray.append(textoLargoArray[j + 1])

                if textoLargoArray[j] == "CUPS GAS":
                    cupsGasArray.append(textoLargoArray[j + 1])

                if j > 0:
                    if len(cupsLuzArray) > tamanoArrayCupLuz:
                        cupsGasArray.append(" ")

        except NoSuchElementException:
            driver.refresh()
            continue

        time.sleep(5)

        try:
            driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div/div[2]/div[2]/div[2]/button").click()
        except NoSuchElementException:
            driver.refresh()
            continue

        time.sleep(5)

        try:
            nombre = driver.find_element(By.XPATH,
                                         "/html/body/div[1]/section/div[2]/div/div/div[3]/div[1]/div[2]/div[1]/div/div/div/div/form/div[1]/div/div[4]/div/div/input")
        except NoSuchElementException:
            driver.refresh()
            continue

        primer_apellido = driver.find_element(By.XPATH, "//input[@id='lastName']")
        segundo_apellido = driver.find_element(By.XPATH, "//input[@id='secondLastName']")
        correo = driver.find_element(By.XPATH, "//input[@id='email']")
        telefono = driver.find_element(By.XPATH, "//input[@id='phone']")

        text_nombre = nombre.get_attribute('value')
        text_Papellido = primer_apellido.get_attribute('value')
        text_Sapellido = segundo_apellido.get_attribute('value')
        text_correo = correo.get_attribute('value')
        text_telefono = telefono.get_attribute('value')

        nombreArray.append(text_nombre)
        primerApellidoArray.append(text_Papellido)
        segundoApellidoArray.append(text_Sapellido)
        correoArray.append(text_correo)
        telefonoArray.append(text_telefono)

        # DATOS DE CUENTA BANCARIA
        time.sleep(10)

        driver.find_element(By.XPATH,
                            "//body[1]/div[1]/section[1]/div[2]/div[1]/div[1]/div[3]/div[1]/div[2]/div[5]/span[1]/span[1]/span[1]").click()

        time.sleep(5)
        driver.find_element(By.XPATH, "//form[@autocomplete='off']//button[@type='submit']").click()

        try:
            cuentaBancaria = driver.find_element(By.XPATH,
                                                 "//body/div[@id='root']/section[@role='main']/div/div/div/div/div/div/div/div/div[2]").text

            cuentaBancariaArray.append(cuentaBancaria[16:len(cuentaBancaria)])
        except NoSuchElementException:
            print("No hay cuenta bancaria")

        for a, value in enumerate(nombreArray):
            worksheet.cell(row=i + 2, column=2, value=value)
        for a, value in enumerate(primerApellidoArray):
            worksheet.cell(row=i + 2, column=3, value=value)
        for a, value in enumerate(segundoApellidoArray):
            worksheet.cell(row=i + 2, column=4, value=value)
        for a, value in enumerate(correoArray):
            worksheet.cell(row=i + 2, column=5, value=value)
        for a, value in enumerate(telefonoArray):
            worksheet.cell(row=i + 2, column=6, value=value)
        for a, value in enumerate(cuentaBancariaArray):
            worksheet.cell(row=i + 2, column=7, value=value)
        for a, value in enumerate(direccionArray):
            worksheet.cell(row=i + 2, column=8, value=value)
        for a, value in enumerate(cupsLuzArray):
            worksheet.cell(row=i + 2, column=9, value=value)
        for a, value in enumerate(cupsGasArray):
            worksheet.cell(row=i + 2, column=10, value=value)

        print(cupsLuzArray)
        print(cupsGasArray)

        workbook.save(filename)
        driver.refresh()


url = "https://checkout.naturgy.es/login"
driver = webdriver.Chrome()
driver.maximize_window()
driver.get(url)
time.sleep(20)
tarea()
